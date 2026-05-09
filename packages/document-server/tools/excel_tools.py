"""
Excel tools
───────────
read_excel  — xlsx/xls → structured JSON (sheets, rows, summary)
data_to_excel — JSON data → xlsx (base64)
"""
from __future__ import annotations

import io
import json
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import xlsxwriter

from config import ALLOWED_EXCEL_MIMES
from utils.file_utils import decode_b64, encode_b64, load_and_validate, temp_file


# ─────────────────────────────────────────────────────────────────────────────
# read_excel
# ─────────────────────────────────────────────────────────────────────────────

def read_excel(file_content: str, sheet_name: str | None = None) -> dict[str, Any]:
    """
    Parse an xlsx/xls file (base64-encoded) and return structured data.

    Parameters
    ----------
    file_content : str
        Base64-encoded xlsx file.
    sheet_name : str | None
        If provided, read only that sheet. Otherwise read all sheets.

    Returns
    -------
    dict with keys:
        filename_hint, sheets (list of sheet dicts), summary
    """
    raw, _ = load_and_validate(file_content, ALLOWED_EXCEL_MIMES, "Excel file", ".xlsx")

    with temp_file(raw, suffix=".xlsx") as path:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

        target_sheets = (
            [sheet_name] if sheet_name and sheet_name in wb.sheetnames
            else wb.sheetnames
        )

        sheets_data: list[dict] = []
        total_rows = 0
        total_cells = 0

        for sname in target_sheets:
            ws = wb[sname]
            rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                # Strip trailing None cells
                row_list = list(row)
                while row_list and row_list[-1] is None:
                    row_list.pop()
                if any(v is not None for v in row_list):
                    rows.append([str(v) if v is not None else "" for v in row_list])

            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []
            total_rows += len(data_rows)
            total_cells += sum(len(r) for r in rows)

            sheets_data.append({
                "sheet_name": sname,
                "headers": headers,
                "row_count": len(data_rows),
                "column_count": len(headers),
                "rows": data_rows,
            })

        wb.close()

    return {
        "sheet_count": len(sheets_data),
        "sheets": sheets_data,
        "summary": {
            "total_sheets": len(target_sheets),
            "total_data_rows": total_rows,
            "total_cells": total_cells,
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# data_to_excel
# ─────────────────────────────────────────────────────────────────────────────

def data_to_excel(
    data: list[dict[str, Any]] | list[list[Any]],
    sheet_name: str = "Sheet1",
    title: str | None = None,
    include_headers: bool = True,
    format_as_table: bool = True,
) -> dict[str, str]:
    """
    Convert a list of dicts (or list of lists) → xlsx file (base64).

    Parameters
    ----------
    data : list[dict] | list[list]
        Rows to write. If list of dicts, keys become column headers.
    sheet_name : str
        Name of the worksheet.
    title : str | None
        Optional bold title written in cell A1 (data starts at row 3).
    include_headers : bool
        Write column headers (default True).
    format_as_table : bool
        Apply alternating row colours and bold headers (default True).

    Returns
    -------
    {"file_content": "<base64>", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
     "row_count": N, "column_count": M}
    """
    if not data:
        raise ValueError("data must not be empty.")

    buffer = io.BytesIO()

    # Normalise to list-of-lists + headers
    if isinstance(data[0], dict):
        headers = list(data[0].keys())
        rows = [[row.get(h, "") for h in headers] for row in data]
    else:
        if include_headers and len(data) > 1:
            headers = [str(v) for v in data[0]]
            rows = data[1:]
        else:
            headers = [get_column_letter(i + 1) for i in range(len(data[0]))]
            rows = data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    start_row = 1

    # Optional title
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
        start_row = 3

    # Headers
    if include_headers:
        header_fill = PatternFill("solid", fgColor="2347E8")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            if format_as_table:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
        start_row += 1

    # Data rows
    alt_fill = PatternFill("solid", fgColor="EEF2FF")
    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=start_row + row_idx - 1, column=col_idx, value=value)
            if format_as_table and row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        col_values = [str(header)] + [str(r[col_idx - 1]) for r in rows if col_idx - 1 < len(r)]
        max_len = min(max((len(v) for v in col_values), default=8), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(buffer)
    raw = buffer.getvalue()

    return {
        "file_content": encode_b64(raw),
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "row_count": len(rows),
        "column_count": len(headers),
        "size_bytes": len(raw),
    }
